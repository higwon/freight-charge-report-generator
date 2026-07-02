from __future__ import annotations

from copy import copy
from dataclasses import dataclass

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .config import BusinessRules


@dataclass
class SheetStyleTemplate:
    font: Font
    fill: PatternFill
    border: Border
    alignment: Alignment
    number_format: str


@dataclass
class RowStyleTemplate:
    first: SheetStyleTemplate
    second: SheetStyleTemplate


@dataclass
class ReportStyle:
    column_widths: dict[str, float]
    row_templates: dict[str, RowStyleTemplate]
    month_format: str
    amount_format: str


class Formatter:
    def __init__(self, rules: BusinessRules | None = None) -> None:
        self.rules = rules or BusinessRules()

    def default_style(self) -> ReportStyle:
        return self._default_style()

    def apply_sheet_layout(self, worksheet: Worksheet, style: ReportStyle) -> None:
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A4"
        for column, width in style.column_widths.items():
            worksheet.column_dimensions[column].width = width

    def style_row(self, worksheet: Worksheet, row: int, kind: str, style: ReportStyle) -> None:
        template = style.row_templates.get(kind) or style.row_templates["body"]
        for column, cell_template in ((1, template.first), (2, template.second)):
            cell = worksheet.cell(row=row, column=column)
            cell.font = copy(cell_template.font)
            cell.fill = copy(cell_template.fill)
            cell.border = copy(cell_template.border)
            cell.alignment = copy(cell_template.alignment)
            cell.number_format = style.amount_format if column == 2 else cell_template.number_format

    def style_raw_sheet(self, worksheet: Worksheet) -> None:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        header_font = Font(name="맑은 고딕", size=10, bold=True, color="1F2937")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        border = Border(bottom=Side(style="thin", color="AAB7C4"))
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        for column_cells in worksheet.columns:
            first = column_cells[0]
            if first.value:
                worksheet.column_dimensions[first.column_letter].width = min(max(len(str(first.value)) + 2, 10), 30)

    def _default_style(self) -> ReportStyle:
        font_name = "맑은 고딕"
        thin_blue = Side(style="thin", color="9EB6CE")
        light_line = Side(style="thin", color="D6E0EA")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        total_fill = PatternFill("solid", fgColor="EAF3F8")
        port_fill = PatternFill("solid", fgColor="F6FAFD")
        no_fill = PatternFill(fill_type=None)

        base_font = Font(name=font_name, size=11, color="111827")
        bold_font = Font(name=font_name, size=11, bold=True, color="111827")
        title_font = Font(name=font_name, size=11, bold=True, color="111827")
        header_font = Font(name=font_name, size=11, bold=True, color="1F2937")

        plain_border = Border()
        header_border = Border(top=thin_blue, bottom=thin_blue)
        separator_border = Border(bottom=light_line)
        total_border = Border(top=thin_blue, bottom=thin_blue)

        def cell(
            *,
            font: Font = base_font,
            fill: PatternFill = no_fill,
            border: Border = plain_border,
            horizontal: str | None = None,
            indent: int = 0,
            number_format: str = "General",
        ) -> SheetStyleTemplate:
            return SheetStyleTemplate(
                font=font,
                fill=fill,
                border=border,
                alignment=Alignment(horizontal=horizontal, vertical="center", indent=indent),
                number_format=number_format,
            )

        amount = self.rules.amount_format
        right_amount = cell(horizontal="right", number_format=amount)

        return ReportStyle(
            column_widths={"A": 59.125, "B": 16.5},
            month_format=self.rules.default_month_format,
            amount_format=amount,
            row_templates={
                "title": RowStyleTemplate(
                    cell(font=title_font, horizontal="left"),
                    cell(font=title_font, horizontal="left"),
                ),
                "blank": RowStyleTemplate(cell(), cell()),
                "header": RowStyleTemplate(
                    cell(font=header_font, fill=header_fill, border=header_border, horizontal="left"),
                    cell(font=header_font, fill=header_fill, border=header_border, horizontal="right", number_format=amount),
                ),
                "month": RowStyleTemplate(
                    cell(font=bold_font, fill=total_fill, border=separator_border, horizontal="left"),
                    cell(font=bold_font, fill=total_fill, border=separator_border, horizontal="right", number_format=amount),
                ),
                "port": RowStyleTemplate(
                    cell(font=bold_font, fill=port_fill, border=separator_border, horizontal="left", indent=1),
                    cell(font=bold_font, fill=port_fill, border=separator_border, horizontal="right", number_format=amount),
                ),
                "customer": RowStyleTemplate(
                    cell(border=separator_border, horizontal="left", indent=2),
                    cell(border=separator_border, horizontal="right", number_format=amount),
                ),
                "grand_total": RowStyleTemplate(
                    cell(font=bold_font, fill=header_fill, border=total_border, horizontal="left"),
                    cell(font=bold_font, fill=header_fill, border=total_border, horizontal="right", number_format=amount),
                ),
                "body": RowStyleTemplate(cell(), right_amount),
            },
        )
