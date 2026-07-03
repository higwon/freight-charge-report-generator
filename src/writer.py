from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError

from .config import ALL_SHEET_NAME, SOURCE_DATA_SHEET_NAME
from .exceptions import OutputPermissionError
from .formatter import Formatter, ReportStyle
from .models import ArApFuncCodeSummary, ArApMonthlyAmount, ArApSummaryRow, FuncCodeSummary, ReportFormat, WorkbookData
from .utils import clean_sheet_title


class ReportWriter:
    def __init__(self, formatter: Formatter | None = None) -> None:
        self.formatter = formatter or Formatter()

    def write(
        self,
        output_path: Path,
        source_data: WorkbookData,
        summaries: list[FuncCodeSummary],
        style: ReportStyle,
        report_format: ReportFormat = ReportFormat.CLASSIC,
    ) -> None:
        workbook = Workbook()
        if report_format == ReportFormat.AR_AP_MONTHLY:
            self._write_ar_ap_monthly_workbook(workbook, source_data, summaries, style)
        elif report_format == ReportFormat.ANALYTIC:
            self._write_analytic_workbook(workbook, source_data, summaries, style)
        else:
            self._write_classic_workbook(workbook, source_data, summaries, style)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            workbook.save(output_path)
        except PermissionError as exc:
            raise OutputPermissionError(
                "출력 파일을 저장할 수 없습니다. 결과 파일이 열려 있다면 닫은 뒤 다시 시도하세요."
            ) from exc

    def _write_classic_workbook(
        self, workbook: Workbook, source_data: WorkbookData, summaries: list[FuncCodeSummary], style: ReportStyle
    ) -> None:
        source_sheet = workbook.active
        source_sheet.title = SOURCE_DATA_SHEET_NAME
        self._write_raw_data(source_sheet, source_data)
        self.formatter.style_raw_sheet(source_sheet)
        for summary in summaries:
            sheet = workbook.create_sheet(clean_sheet_title(summary.func_code))
            self._write_summary_sheet(sheet, summary, style)

    def _write_analytic_workbook(
        self, workbook: Workbook, source_data: WorkbookData, summaries: list[FuncCodeSummary], style: ReportStyle
    ) -> None:
        func_summaries = [summary for summary in summaries if summary.func_code != ALL_SHEET_NAME]
        source_sheet = workbook.active
        source_sheet.title = SOURCE_DATA_SHEET_NAME
        self._write_raw_data(source_sheet, source_data)
        self.formatter.style_raw_sheet(source_sheet, enable_filter=True)

        overview = workbook.create_sheet("Overview")
        self._write_overview(overview, source_data, func_summaries, style)

        all_sheet = workbook.create_sheet(ALL_SHEET_NAME)
        self._write_analytic_all(all_sheet, func_summaries, style)
        for summary in func_summaries:
            sheet = workbook.create_sheet(clean_sheet_title(summary.func_code))
            self._write_summary_sheet(sheet, summary, style)

    def _write_ar_ap_monthly_workbook(
        self,
        workbook: Workbook,
        source_data: WorkbookData,
        summaries: list[ArApFuncCodeSummary],
        style: ReportStyle,
    ) -> None:
        source_sheet = workbook.active
        source_sheet.title = SOURCE_DATA_SHEET_NAME
        self._write_raw_data(source_sheet, source_data)
        self.formatter.style_raw_sheet(source_sheet, enable_filter=True)

        overview = workbook.create_sheet("Overview")
        self._write_ar_ap_overview(overview, source_data, summaries, style)

        for summary in summaries:
            sheet = workbook.create_sheet(clean_sheet_title(summary.func_code))
            self._write_ar_ap_monthly_sheet(sheet, summary, style)

    def _write_ar_ap_monthly_sheet(
        self, sheet, summary: ArApFuncCodeSummary, style: ReportStyle
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "H12"
        sheet.sheet_properties.outlinePr.summaryBelow = False

        key_headers = [summary.port_label, "거래처코드", "거래처명"] if summary.port_label else ["거래처코드", "거래처명"]
        key_headers = [header for header in key_headers if header]
        key_count = len(key_headers)
        detail_start_row = 10
        first_data_row = detail_start_row + 2
        cumulative_headers = ["누적 매출계", "누적 매입계", "누적 차이", "누적 마진율"]
        max_column = key_count + len(cumulative_headers) + len(summary.months) * 4
        rank_first_column = 4 + len(summary.months)
        rank_last_column = rank_first_column + (10 if summary.port_label else 4)
        max_column = max(max_column, rank_last_column)

        if max_column >= 1:
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
        sheet.cell(1, 1, f"{summary.func_code} {summary.category} AR/AP 월별 보고서")
        sheet.cell(2, 1, "기준")
        sheet.cell(2, 2, f"{summary.port_label} + 거래처" if summary.port_label else "거래처")

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        month_fill = PatternFill("solid", fgColor="EAF3F8")
        diff_fill = PatternFill("solid", fgColor="FFF2CC")
        summary_fill = PatternFill("solid", fgColor="E2F0D9")
        title_fill = PatternFill("solid", fgColor="1F4E78")
        border = Border(
            left=Side(style="thin", color="D6E0EA"),
            right=Side(style="thin", color="D6E0EA"),
            top=Side(style="thin", color="D6E0EA"),
            bottom=Side(style="thin", color="D6E0EA"),
        )
        header_font = Font(name="Malgun Gothic", size=10, bold=True, color="1F2937")

        title = sheet.cell(1, 1)
        title.font = Font(name="Malgun Gothic", size=14, bold=True, color="FFFFFF")
        title.fill = title_fill
        title.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 26

        summary_columns = [2] + list(range(3, 3 + len(summary.months)))
        sheet.cell(4, 1, "구분")
        sheet.cell(4, 2, "총 누적")
        for offset, month in enumerate(summary.months, start=3):
            sheet.cell(4, offset, month)
        for row_index, label in enumerate(("매출계", "매입계", "차이", "마진율"), start=5):
            sheet.cell(row_index, 1, label)
        total_by_month = self._ar_ap_month_totals(summary)
        grand_total = self._ar_ap_grand_total(summary)
        for col in summary_columns:
            month = None if col == 2 else summary.months[col - 3]
            amount = grand_total if month is None else total_by_month[month]
            ar_cell = sheet.cell(5, col, self._numeric(amount.ar))
            ap_cell = sheet.cell(6, col, self._numeric(amount.ap))
            diff_cell = sheet.cell(7, col, f"={ar_cell.coordinate}-{ap_cell.coordinate}")
            margin_cell = sheet.cell(8, col)
            margin_cell.value = self._margin_formula(ar_cell.coordinate, ap_cell.coordinate, diff_cell.coordinate)
            for cell in (ar_cell, ap_cell, diff_cell):
                cell.number_format = style.amount_format
                cell.alignment = Alignment(horizontal="right", vertical="center")
            margin_cell.number_format = "0.0%"
            margin_cell.alignment = Alignment(horizontal="right", vertical="center")
        for row in range(4, 9):
            for col in range(1, 3 + len(summary.months)):
                cell = sheet.cell(row, col)
                cell.fill = summary_fill if row == 4 or col == 1 else PatternFill(fill_type=None)
                cell.font = header_font if row == 4 or col == 1 else Font(name="Malgun Gothic", size=10, color="111827")
                cell.border = border

        rank_start_col = rank_first_column
        rank_tables = self._ar_ap_rankings(summary)
        if summary.port_label:
            self._write_amount_rank_table(
                sheet, 4, rank_start_col, "매출 상위 Port", rank_tables["port_ar"], style
            )
            self._write_amount_rank_table(
                sheet, 4, rank_start_col + 3, "매입 상위 Port", rank_tables["port_ap"], style
            )
            rank_start_col += 6
        self._write_amount_rank_table(
            sheet, 4, rank_start_col, "매출 상위 Customer", rank_tables["customer_ar"], style
        )
        self._write_amount_rank_table(
            sheet, 4, rank_start_col + 3, "매입 상위 Customer", rank_tables["customer_ap"], style
        )

        sheet.merge_cells(
            start_row=detail_start_row,
            start_column=1,
            end_row=detail_start_row,
            end_column=key_count,
        )
        key_group_cell = sheet.cell(detail_start_row, 1, "기본정보")
        key_group_cell.fill = header_fill
        key_group_cell.font = header_font
        key_group_cell.alignment = Alignment(horizontal="center", vertical="center")
        key_group_cell.border = border
        for column, header in enumerate(key_headers, start=1):
            sub_cell = sheet.cell(detail_start_row + 1, column, header)
            sub_cell.fill = header_fill
            sub_cell.font = header_font
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")
            sub_cell.border = border

        column = key_count + 1
        sheet.merge_cells(
            start_row=detail_start_row,
            start_column=column,
            end_row=detail_start_row,
            end_column=column + len(cumulative_headers) - 1,
        )
        cumulative_group_cell = sheet.cell(detail_start_row, column, "누적 합계")
        cumulative_group_cell.fill = summary_fill
        cumulative_group_cell.font = header_font
        cumulative_group_cell.alignment = Alignment(horizontal="center", vertical="center")
        cumulative_group_cell.border = border
        for header in cumulative_headers:
            sub_cell = sheet.cell(detail_start_row + 1, column, header)
            sub_cell.fill = summary_fill
            sub_cell.font = header_font
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")
            sub_cell.border = border
            column += 1

        for month in summary.months:
            sheet.merge_cells(start_row=detail_start_row, start_column=column, end_row=detail_start_row, end_column=column + 3)
            month_cell = sheet.cell(detail_start_row, column, month)
            month_cell.fill = month_fill
            month_cell.font = header_font
            month_cell.alignment = Alignment(horizontal="center", vertical="center")
            month_cell.border = border
            for offset, label in enumerate(("AR계", "AP계", "차이", "마진율")):
                cell = sheet.cell(detail_start_row + 1, column + offset, label)
                cell.fill = diff_fill if label in {"차이", "마진율"} else header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            column += 4

        for row_index, row_summary in enumerate(summary.rows, start=first_data_row):
            values = []
            if summary.port_label:
                values.append(row_summary.port or "(Blank)")
            values.extend([row_summary.customer_code, row_summary.customer_name])
            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row_index, column_index, value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

            column = key_count + 1
            total_ar = sum((row_summary.monthly[month].ar for month in summary.months), Decimal("0"))
            total_ap = sum((row_summary.monthly[month].ap for month in summary.months), Decimal("0"))
            ar_cell = sheet.cell(row_index, column, self._numeric(total_ar))
            ap_cell = sheet.cell(row_index, column + 1, self._numeric(total_ap))
            diff_cell = sheet.cell(row_index, column + 2, f"={ar_cell.coordinate}-{ap_cell.coordinate}")
            margin_cell = sheet.cell(row_index, column + 3)
            margin_cell.value = self._margin_formula(ar_cell.coordinate, ap_cell.coordinate, diff_cell.coordinate)
            for cell in (ar_cell, ap_cell, diff_cell):
                cell.number_format = style.amount_format
                cell.border = border
                cell.alignment = Alignment(horizontal="right", vertical="center")
            margin_cell.number_format = "0.0%"
            margin_cell.border = border
            margin_cell.alignment = Alignment(horizontal="right", vertical="center")
            column += 4

            for month in summary.months:
                amount = row_summary.monthly[month]
                values = (amount.ar, amount.ap)
                for offset, number in enumerate(values):
                    cell = sheet.cell(row_index, column + offset, self._numeric(number))
                    cell.number_format = style.amount_format
                    cell.border = border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                ar_ref = sheet.cell(row_index, column).coordinate
                ap_ref = sheet.cell(row_index, column + 1).coordinate
                diff_ref = sheet.cell(row_index, column + 2).coordinate
                diff_cell = sheet.cell(row_index, column + 2, f"={ar_ref}-{ap_ref}")
                diff_cell.number_format = style.amount_format
                diff_cell.border = border
                diff_cell.alignment = Alignment(horizontal="right", vertical="center")
                margin_cell = sheet.cell(row_index, column + 3)
                margin_cell.value = self._margin_formula(ar_ref, ap_ref, diff_ref)
                margin_cell.number_format = "0.0%"
                margin_cell.alignment = Alignment(horizontal="right", vertical="center")
                margin_cell.border = border
                column += 4

        if sheet.max_row >= detail_start_row + 1 and max_column >= 1:
            sheet.auto_filter.ref = f"A{detail_start_row + 1}:{get_column_letter(max_column)}{sheet.max_row}"

        widths = [13, 14, 38] if summary.port_label else [14, 38]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for column_index in range(key_count + 1, key_count + 5):
            sheet.column_dimensions[get_column_letter(column_index)].width = 14
        for column_index in range(key_count + 1, max_column + 1):
            sheet.column_dimensions[get_column_letter(column_index)].width = max(
                sheet.column_dimensions[get_column_letter(column_index)].width or 0, 13
            )
        for month_index in range(len(summary.months)):
            start_col = key_count + 5 + month_index * 4
            for col in range(start_col, start_col + 4):
                sheet.column_dimensions[get_column_letter(col)].outlineLevel = 1

    def _ar_ap_rankings(self, summary: ArApFuncCodeSummary) -> dict[str, list[tuple[str, Decimal]]]:
        port_ar: dict[str, Decimal] = defaultdict(Decimal)
        port_ap: dict[str, Decimal] = defaultdict(Decimal)
        customer_ar: dict[str, Decimal] = defaultdict(Decimal)
        customer_ap: dict[str, Decimal] = defaultdict(Decimal)
        for row in summary.rows:
            total = self._ar_ap_row_total(row, summary.months)
            if row.port:
                port_ar[row.port] += total.ar
                port_ap[row.port] += total.ap
            customer_label = row.customer_name or row.customer_code or "(Blank)"
            customer_ar[customer_label] += total.ar
            customer_ap[customer_label] += total.ap

        def top(values: dict[str, Decimal], limit: int = 5) -> list[tuple[str, Decimal]]:
            return sorted(values.items(), key=lambda item: item[1], reverse=True)[:limit]

        return {
            "port_ar": top(port_ar),
            "port_ap": top(port_ap),
            "customer_ar": top(customer_ar),
            "customer_ap": top(customer_ap),
        }

    def _write_amount_rank_table(
        self, sheet, start_row: int, start_col: int, title: str, rows, style: ReportStyle
    ) -> int:
        sheet.cell(start_row, start_col, title)
        sheet.cell(start_row, start_col + 1, "금액")
        self._style_table_header(sheet, start_row, start_col, start_col + 1)
        for offset, (label, amount) in enumerate(rows, start=1):
            row = start_row + offset
            sheet.cell(row, start_col, label)
            amount_cell = sheet.cell(row, start_col + 1, self._numeric(amount))
            amount_cell.number_format = style.amount_format
            amount_cell.alignment = Alignment(horizontal="right", vertical="center")
        sheet.column_dimensions[get_column_letter(start_col)].width = max(
            sheet.column_dimensions[get_column_letter(start_col)].width or 0, 22
        )
        sheet.column_dimensions[get_column_letter(start_col + 1)].width = max(
            sheet.column_dimensions[get_column_letter(start_col + 1)].width or 0, 14
        )
        return start_row + len(rows)

    def _write_ar_ap_overview_legacy(
        self,
        sheet,
        source_data: WorkbookData,
        summaries: list[ArApFuncCodeSummary],
        style: ReportStyle,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A6"
        sheet.merge_cells("A1:K1")
        sheet["A1"] = "AR/AP Monthly Overview"
        sheet["A1"].font = Font(name="Malgun Gothic", size=18, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 30

        grand_total = self._ar_ap_grand_total_all(summaries)
        metadata = (
            ("A3", "원본 시트", "B3", source_data.source_sheet_name),
            ("D3", "Source 행 수", "E3", source_data.record_count),
            ("G3", "총 차이", "H3", self._numeric(grand_total.difference)),
            ("J3", "생성 일시", "K3", datetime.now().strftime("%Y-%m-%d %H:%M")),
        )
        for label_cell, label, value_cell, value in metadata:
            sheet[label_cell] = label
            sheet[label_cell].font = Font(name="Malgun Gothic", bold=True, color="44546A")
            sheet[value_cell] = value
        sheet["H3"].number_format = style.amount_format

        kpi_rows = [
            ("총 매출계", grand_total.ar),
            ("총 매입계", grand_total.ap),
            ("총 차이", grand_total.difference),
            ("전체 마진율", self._margin_value(grand_total.ar, grand_total.ap)),
            ("매출만발생 건수", self._count_margin_status(summaries, "매출만발생")),
            ("매입만발생 건수", self._count_margin_status(summaries, "매입만발생")),
            ("마진 음수 건수", self._count_negative_margin(summaries)),
        ]
        self._write_ar_ap_kpi_table(sheet, 5, 1, "전체 KPI", kpi_rows, style)

        code_rows = []
        for summary in summaries:
            total = self._ar_ap_grand_total(summary)
            code_rows.append(
                (
                    summary.func_code,
                    summary.category,
                    total.ar,
                    total.ap,
                    total.difference,
                    self._margin_value(total.ar, total.ap),
                    len(summary.rows),
                    len({row.port for row in summary.rows if row.port}),
                    self._count_margin_status([summary], "매출만발생"),
                    self._count_margin_status([summary], "매입만발생"),
                )
            )
        self._write_ar_ap_code_table(sheet, 5, 4, code_rows, style)

        month_rows = []
        for month in sorted({month for summary in summaries for month in summary.months}):
            total = self._ar_ap_month_total_all(summaries, month)
            month_rows.append((month, total.ar, total.ap, total.difference, self._margin_value(total.ar, total.ap)))
        month_end = self._write_ar_ap_month_table(sheet, 5, 15, month_rows, style)

        ports: dict[str, Decimal] = defaultdict(Decimal)
        customers: dict[str, Decimal] = defaultdict(Decimal)
        review_rows = []
        customer_totals: dict[tuple[str, str], ArApMonthlyAmount] = {}
        for summary in summaries:
            for row in summary.rows:
                row_total = self._ar_ap_row_total(row, summary.months)
                if row.port:
                    ports[row.port] += row_total.difference
                customer_key = (row.customer_code, row.customer_name)
                current = customer_totals.get(customer_key, ArApMonthlyAmount())
                customer_totals[customer_key] = ArApMonthlyAmount(
                    ar=current.ar + row_total.ar,
                    ap=current.ap + row_total.ap,
                )
                customers[row.customer_name] += row_total.difference
                status = self._margin_value(row_total.ar, row_total.ap)
                if status in {"매출만발생", "매입만발생"} or row_total.difference < 0:
                    review_rows.append(
                        (
                            summary.func_code,
                            summary.category,
                            row.port or "",
                            row.customer_code,
                            row.customer_name,
                            row_total.ar,
                            row_total.ap,
                            row_total.difference,
                            status,
                        )
                    )

        self._write_overview_table(
            sheet,
            17,
            1,
            "상위 Port",
            sorted(ports.items(), key=lambda item: item[1], reverse=True)[:10],
            style,
        )
        self._write_overview_table(
            sheet,
            17,
            4,
            "상위 Customer",
            sorted(customers.items(), key=lambda item: item[1], reverse=True)[:10],
            style,
        )
        self._write_ar_ap_review_table(sheet, 17, 7, review_rows[:20], style)

        ranked_customers = [
            (code, name, total.ar, total.ap, total.difference, self._margin_value(total.ar, total.ap))
            for (code, name), total in customer_totals.items()
        ]
        ranked_customers.sort(key=lambda item: item[4], reverse=True)
        self._write_ar_ap_rank_table(sheet, 42, 1, "차이 상위 거래처 Top 10", ranked_customers[:10], style)
        self._write_ar_ap_rank_table(sheet, 42, 8, "차이 하위 거래처 Bottom 10", list(reversed(ranked_customers[-10:])), style)

        if code_rows:
            chart = BarChart()
            chart.title = "Func Code별 차이"
            chart.y_axis.title = "차이"
            chart.add_data(Reference(sheet, min_col=8, min_row=5, max_row=5 + len(code_rows)), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=4, min_row=6, max_row=5 + len(code_rows)))
            chart.height = 7
            chart.width = 12
            sheet.add_chart(chart, "A58")
        if month_end >= 6:
            chart = LineChart()
            chart.title = "월별 차이 추이"
            chart.y_axis.title = "차이"
            chart.add_data(Reference(sheet, min_col=18, min_row=5, max_row=month_end), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=15, min_row=6, max_row=month_end))
            chart.height = 7
            chart.width = 12
            sheet.add_chart(chart, "H58")

        for column_index in range(1, 21):
            sheet.column_dimensions[get_column_letter(column_index)].width = 15
        sheet.column_dimensions["E"].width = 22
        sheet.column_dimensions["K"].width = 34
        sheet.column_dimensions["L"].width = 14

    def _write_ar_ap_overview(
        self,
        sheet,
        source_data: WorkbookData,
        summaries: list[ArApFuncCodeSummary],
        style: ReportStyle,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A8"
        sheet.merge_cells("A1:L1")
        sheet["A1"] = "AR/AP Monthly Overview"
        sheet["A1"].font = Font(name="Malgun Gothic", size=18, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 30

        grand_total = self._ar_ap_grand_total_all(summaries)
        margin = self._margin_value(grand_total.ar, grand_total.ap)
        self._write_overview_meta(sheet, source_data, grand_total, style)
        self._write_overview_cards(
            sheet,
            4,
            [
                ("총 매출계", grand_total.ar),
                ("총 매입계", grand_total.ap),
                ("총 차이", grand_total.difference),
                ("전체 마진율", margin),
            ],
            style,
        )

        code_rows = []
        review_rows = []
        port_ar: dict[str, Decimal] = defaultdict(Decimal)
        port_ap: dict[str, Decimal] = defaultdict(Decimal)
        customer_ar: dict[str, Decimal] = defaultdict(Decimal)
        customer_ap: dict[str, Decimal] = defaultdict(Decimal)
        for summary in summaries:
            total = self._ar_ap_grand_total(summary)
            code_rows.append(
                (
                    summary.func_code,
                    summary.category,
                    total.ar,
                    total.ap,
                    total.difference,
                    self._margin_value(total.ar, total.ap),
                    len(summary.rows),
                    self._count_margin_status([summary], "매출만발생"),
                    self._count_margin_status([summary], "매입만발생"),
                )
            )
            for row in summary.rows:
                row_total = self._ar_ap_row_total(row, summary.months)
                if row.port:
                    port_ar[row.port] += row_total.ar
                    port_ap[row.port] += row_total.ap
                customer_label = row.customer_name or row.customer_code or "(Blank)"
                customer_ar[customer_label] += row_total.ar
                customer_ap[customer_label] += row_total.ap
                status = self._margin_value(row_total.ar, row_total.ap)
                if status in {"매출만발생", "매입만발생"} or row_total.difference < 0:
                    review_rows.append(
                        (
                            summary.func_code,
                            row.port or "",
                            row.customer_name,
                            row_total.ar,
                            row_total.ap,
                            row_total.difference,
                            status,
                        )
                    )

        code_end = self._write_ar_ap_compact_code_table(sheet, 8, 1, code_rows, style)

        month_rows = []
        for month in sorted({month for summary in summaries for month in summary.months}):
            total = self._ar_ap_month_total_all(summaries, month)
            month_rows.append((month, total.ar, total.ap, total.difference, self._margin_value(total.ar, total.ap)))
        month_end = self._write_ar_ap_month_table(sheet, 8, 11, month_rows, style)

        secondary_start = max(code_end, month_end) + 3
        review_rows.sort(key=lambda item: item[5])
        self._write_ar_ap_compact_review_table(sheet, secondary_start, 1, review_rows[:12], style)

        top_start = secondary_start
        self._write_overview_table(
            sheet,
            top_start,
            10,
            "매출 상위 Port",
            sorted(port_ar.items(), key=lambda item: item[1], reverse=True)[:10],
            style,
        )
        self._write_overview_table(
            sheet,
            top_start,
            13,
            "매입 상위 Port",
            sorted(port_ap.items(), key=lambda item: item[1], reverse=True)[:10],
            style,
        )
        customer_top_start = top_start + 12
        self._write_overview_table(
            sheet,
            customer_top_start,
            10,
            "매출 상위 Customer",
            sorted(customer_ar.items(), key=lambda item: item[1], reverse=True)[:10],
            style,
        )
        self._write_overview_table(
            sheet,
            customer_top_start,
            13,
            "매입 상위 Customer",
            sorted(customer_ap.items(), key=lambda item: item[1], reverse=True)[:10],
            style,
        )
        for header_col in (11, 14):
            sheet.cell(top_start, header_col, "금액")
            sheet.cell(customer_top_start, header_col, "금액")

        if month_end >= 9:
            chart = LineChart()
            chart.title = "월별 차이 추이"
            chart.y_axis.title = "차이"
            chart.add_data(Reference(sheet, min_col=14, min_row=8, max_row=month_end), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=11, min_row=9, max_row=month_end))
            chart.height = 7
            chart.width = 13
            sheet.add_chart(chart, f"A{secondary_start + 25}")

        widths = {
            "A": 12,
            "B": 12,
            "C": 14,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 12,
            "H": 12,
            "I": 12,
            "J": 3,
            "K": 12,
            "L": 14,
            "M": 14,
            "N": 14,
            "O": 14,
            "P": 3,
            "Q": 18,
            "R": 34,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

    def _write_overview_meta(self, sheet, source_data: WorkbookData, grand_total: ArApMonthlyAmount, style: ReportStyle) -> None:
        metadata = (
            ("A3", "원본 시트", "B3", source_data.source_sheet_name),
            ("D3", "Source 행 수", "E3", source_data.record_count),
            ("G3", "총 차이", "H3", self._numeric(grand_total.difference)),
            ("J3", "생성 일시", "K3", datetime.now().strftime("%Y-%m-%d %H:%M")),
        )
        for label_cell, label, value_cell, value in metadata:
            sheet[label_cell] = label
            sheet[label_cell].font = Font(name="Malgun Gothic", bold=True, color="44546A")
            sheet[value_cell] = value
        sheet["H3"].number_format = style.amount_format

    def _write_overview_cards(self, sheet, row: int, cards, style: ReportStyle) -> None:
        fills = ("D9EAF7", "E2F0D9", "FFF2CC", "EADCF8")
        for index, (label, value) in enumerate(cards):
            col = 1 + index * 3
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            sheet.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
            label_cell = sheet.cell(row, col, label)
            value_cell = sheet.cell(row + 1, col)
            label_cell.fill = PatternFill("solid", fgColor=fills[index % len(fills)])
            label_cell.font = Font(name="Malgun Gothic", bold=True, color="1F2937")
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            self._write_margin_cell(value_cell, value) if label.endswith("마진율") else self._write_amount_or_text(value_cell, value, style)
            value_cell.font = Font(name="Malgun Gothic", size=14, bold=True, color="111827")
            value_cell.fill = PatternFill("solid", fgColor="FFFFFF")
            for merged_row in (row, row + 1):
                for merged_col in (col, col + 1):
                    sheet.cell(merged_row, merged_col).border = Border(
                        left=Side(style="thin", color="D6E0EA"),
                        right=Side(style="thin", color="D6E0EA"),
                        top=Side(style="thin", color="D6E0EA"),
                        bottom=Side(style="thin", color="D6E0EA"),
                    )

    def _write_ar_ap_compact_code_table(self, sheet, start_row: int, start_col: int, rows, style: ReportStyle) -> int:
        headers = ("코드", "구분", "매출계", "매입계", "차이", "마진율", "거래처", "매출만", "매입만")
        for offset, header in enumerate(headers):
            sheet.cell(start_row, start_col + offset, header)
        self._style_table_header(sheet, start_row, start_col, start_col + len(headers) - 1)
        for row_offset, values in enumerate(rows, start=1):
            row = start_row + row_offset
            for col_offset, value in enumerate(values):
                cell = sheet.cell(row, start_col + col_offset)
                if col_offset == 5:
                    self._write_margin_cell(cell, value)
                else:
                    self._write_amount_or_text(cell, value, style)
        return start_row + len(rows)

    def _write_ar_ap_compact_review_table(self, sheet, start_row: int, start_col: int, rows, style: ReportStyle) -> int:
        headers = ("검토 대상", "항구", "거래처명", "매출계", "매입계", "차이", "상태")
        for offset, header in enumerate(headers):
            sheet.cell(start_row, start_col + offset, header)
        self._style_table_header(sheet, start_row, start_col, start_col + len(headers) - 1)
        for row_offset, values in enumerate(rows, start=1):
            row = start_row + row_offset
            for col_offset, value in enumerate(values):
                cell = sheet.cell(row, start_col + col_offset)
                if isinstance(value, Decimal):
                    cell.value = self._numeric(value)
                    cell.number_format = style.amount_format
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        return start_row + len(rows)

    @staticmethod
    def _write_amount_or_text(cell, value, style: ReportStyle) -> None:
        if isinstance(value, Decimal):
            cell.value = ReportWriter._numeric(value)
            cell.number_format = style.amount_format
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _margin_formula(ar_ref: str, ap_ref: str, diff_ref: str) -> str:
        return (
            f'=IF(AND({ar_ref}=0,{ap_ref}=0),"-",'
            f'IF(AND({ar_ref}>0,{ap_ref}=0),"매출만발생",'
            f'IF(AND({ar_ref}=0,{ap_ref}>0),"매입만발생",{diff_ref}/{ar_ref})))'
        )

    @staticmethod
    def _margin_value(ar: Decimal, ap: Decimal) -> Decimal | str:
        difference = ar - ap
        if ar == 0 and ap == 0:
            return "-"
        if ar > 0 and ap == 0:
            return "매출만발생"
        if ar == 0 and ap > 0:
            return "매입만발생"
        return difference / ar

    @staticmethod
    def _ar_ap_row_total(row: ArApSummaryRow, months: list[str]) -> ArApMonthlyAmount:
        return ArApMonthlyAmount(
            ar=sum((row.monthly[month].ar for month in months), Decimal("0")),
            ap=sum((row.monthly[month].ap for month in months), Decimal("0")),
        )

    def _ar_ap_grand_total(self, summary: ArApFuncCodeSummary) -> ArApMonthlyAmount:
        return ArApMonthlyAmount(
            ar=sum((self._ar_ap_row_total(row, summary.months).ar for row in summary.rows), Decimal("0")),
            ap=sum((self._ar_ap_row_total(row, summary.months).ap for row in summary.rows), Decimal("0")),
        )

    def _ar_ap_grand_total_all(self, summaries: list[ArApFuncCodeSummary]) -> ArApMonthlyAmount:
        return ArApMonthlyAmount(
            ar=sum((self._ar_ap_grand_total(summary).ar for summary in summaries), Decimal("0")),
            ap=sum((self._ar_ap_grand_total(summary).ap for summary in summaries), Decimal("0")),
        )

    @staticmethod
    def _ar_ap_month_totals(summary: ArApFuncCodeSummary) -> dict[str, ArApMonthlyAmount]:
        return {
            month: ArApMonthlyAmount(
                ar=sum((row.monthly[month].ar for row in summary.rows), Decimal("0")),
                ap=sum((row.monthly[month].ap for row in summary.rows), Decimal("0")),
            )
            for month in summary.months
        }

    @staticmethod
    def _ar_ap_month_total_all(summaries: list[ArApFuncCodeSummary], month: str) -> ArApMonthlyAmount:
        return ArApMonthlyAmount(
            ar=sum(
                (row.monthly[month].ar for summary in summaries if month in summary.months for row in summary.rows),
                Decimal("0"),
            ),
            ap=sum(
                (row.monthly[month].ap for summary in summaries if month in summary.months for row in summary.rows),
                Decimal("0"),
            ),
        )

    def _count_margin_status(self, summaries: list[ArApFuncCodeSummary], status: str) -> int:
        count = 0
        for summary in summaries:
            for row in summary.rows:
                total = self._ar_ap_row_total(row, summary.months)
                if self._margin_value(total.ar, total.ap) == status:
                    count += 1
        return count

    def _count_negative_margin(self, summaries: list[ArApFuncCodeSummary]) -> int:
        count = 0
        for summary in summaries:
            for row in summary.rows:
                total = self._ar_ap_row_total(row, summary.months)
                if total.ar > 0 and total.difference < 0:
                    count += 1
        return count

    @staticmethod
    def _style_table_header(sheet, row: int, start_col: int, end_col: int) -> None:
        fill = PatternFill("solid", fgColor="D9EAF7")
        border = Border(bottom=Side(style="thin", color="9EB6CE"))
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row, col)
            cell.font = Font(name="Malgun Gothic", bold=True, color="1F2937")
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _write_margin_cell(cell, value: Decimal | str) -> None:
        cell.value = ReportWriter._numeric(value) if isinstance(value, Decimal) else value
        cell.number_format = "0.0%" if isinstance(value, Decimal) else "General"
        cell.alignment = Alignment(horizontal="right" if isinstance(value, Decimal) else "center", vertical="center")

    def _write_ar_ap_kpi_table(self, sheet, start_row: int, start_col: int, title: str, rows, style: ReportStyle) -> int:
        sheet.cell(start_row, start_col, title)
        sheet.cell(start_row, start_col + 1, "값")
        self._style_table_header(sheet, start_row, start_col, start_col + 1)
        for offset, (label, value) in enumerate(rows, start=1):
            row = start_row + offset
            sheet.cell(row, start_col, label)
            value_cell = sheet.cell(row, start_col + 1)
            if isinstance(value, Decimal):
                value_cell.value = self._numeric(value)
                value_cell.number_format = "0.0%" if abs(value) <= 1 and "마진율" in label else style.amount_format
                value_cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                value_cell.value = value
        return start_row + len(rows)

    def _write_ar_ap_code_table(self, sheet, start_row: int, start_col: int, rows, style: ReportStyle) -> int:
        headers = ("Func Code", "구분", "매출계", "매입계", "차이", "마진율", "거래처 수", "항구 수", "매출만", "매입만")
        for offset, header in enumerate(headers):
            sheet.cell(start_row, start_col + offset, header)
        self._style_table_header(sheet, start_row, start_col, start_col + len(headers) - 1)
        for row_offset, values in enumerate(rows, start=1):
            row = start_row + row_offset
            for col_offset, value in enumerate(values):
                cell = sheet.cell(row, start_col + col_offset)
                if col_offset == 5:
                    self._write_margin_cell(cell, value)
                elif isinstance(value, Decimal):
                    cell.value = self._numeric(value)
                    cell.number_format = style.amount_format
                else:
                    cell.value = value
        return start_row + len(rows)

    def _write_ar_ap_month_table(self, sheet, start_row: int, start_col: int, rows, style: ReportStyle) -> int:
        headers = ("월", "매출계", "매입계", "차이", "마진율")
        for offset, header in enumerate(headers):
            sheet.cell(start_row, start_col + offset, header)
        self._style_table_header(sheet, start_row, start_col, start_col + len(headers) - 1)
        for row_offset, values in enumerate(rows, start=1):
            row = start_row + row_offset
            for col_offset, value in enumerate(values):
                cell = sheet.cell(row, start_col + col_offset)
                if col_offset == 4:
                    self._write_margin_cell(cell, value)
                elif isinstance(value, Decimal):
                    cell.value = self._numeric(value)
                    cell.number_format = style.amount_format
                else:
                    cell.value = value
        return start_row + len(rows)

    def _write_ar_ap_review_table(self, sheet, start_row: int, start_col: int, rows, style: ReportStyle) -> int:
        headers = ("검토", "코드", "구분", "항구", "거래처코드", "거래처명", "매출계", "매입계", "차이", "상태")
        sheet.cell(start_row, start_col, "검토 대상")
        self._style_table_header(sheet, start_row, start_col, start_col + len(headers) - 1)
        for offset, header in enumerate(headers[1:], start=1):
            sheet.cell(start_row, start_col + offset, header)
        for row_offset, values in enumerate(rows, start=1):
            row = start_row + row_offset
            for col_offset, value in enumerate(values, start=1):
                cell = sheet.cell(row, start_col + col_offset)
                if isinstance(value, Decimal):
                    cell.value = self._numeric(value)
                    cell.number_format = style.amount_format
                else:
                    cell.value = value
        return start_row + len(rows)

    def _write_ar_ap_rank_table(self, sheet, start_row: int, start_col: int, title: str, rows, style: ReportStyle) -> int:
        headers = (title, "거래처명", "매출계", "매입계", "차이", "마진율")
        for offset, header in enumerate(headers):
            sheet.cell(start_row, start_col + offset, header)
        self._style_table_header(sheet, start_row, start_col, start_col + len(headers) - 1)
        for row_offset, (code, name, ar, ap, difference, margin) in enumerate(rows, start=1):
            row = start_row + row_offset
            values = (code, name, ar, ap, difference, margin)
            for col_offset, value in enumerate(values):
                cell = sheet.cell(row, start_col + col_offset)
                if col_offset == 5:
                    self._write_margin_cell(cell, value)
                elif isinstance(value, Decimal):
                    cell.value = self._numeric(value)
                    cell.number_format = style.amount_format
                else:
                    cell.value = value
        return start_row + len(rows)


    def _write_overview(
        self, sheet, source_data: WorkbookData, summaries: list[FuncCodeSummary], style: ReportStyle
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A5"
        sheet.merge_cells("A1:K1")
        sheet["A1"] = "Freight Charge Overview"
        sheet["A1"].font = Font(name="맑은 고딕", size=18, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 30

        total_amount = sum((summary.amount for summary in summaries), Decimal("0"))
        metadata = (
            ("A3", "원본 시트", "B3", source_data.source_sheet_name),
            ("D3", "Source 행 수", "E3", source_data.record_count),
            ("G3", "총 운임", "H3", self._numeric(total_amount)),
            ("J3", "생성 일시", "K3", datetime.now().strftime("%Y-%m-%d %H:%M")),
        )
        for label_cell, label, value_cell, value in metadata:
            sheet[label_cell] = label
            sheet[label_cell].font = Font(name="맑은 고딕", bold=True, color="44546A")
            sheet[value_cell] = value
        sheet["H3"].number_format = style.amount_format

        monthly: dict[str, Decimal] = defaultdict(Decimal)
        ports: dict[str, Decimal] = defaultdict(Decimal)
        customers: dict[str, Decimal] = defaultdict(Decimal)
        for summary in summaries:
            for month in summary.months:
                monthly[month.label] += month.amount
                for port in month.ports:
                    ports[port.name] += port.amount
                    for customer in port.customers:
                        customers[customer.name] += customer.amount

        func_end = self._write_overview_table(
            sheet, 5, 1, "Func Code", [(s.func_code, s.amount) for s in summaries], style
        )
        month_end = self._write_overview_table(sheet, 5, 4, "월별 합계", sorted(monthly.items()), style)
        self._write_overview_table(
            sheet, 5, 7, "상위 Port", sorted(ports.items(), key=lambda item: item[1], reverse=True)[:10], style
        )
        self._write_overview_table(
            sheet, 5, 10, "상위 Customer", sorted(customers.items(), key=lambda item: item[1], reverse=True)[:10], style
        )

        if func_end >= 6:
            chart = BarChart()
            chart.title = "Func Code별 운임"
            chart.y_axis.title = "Loc Amt"
            chart.add_data(Reference(sheet, min_col=2, min_row=5, max_row=func_end), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=1, min_row=6, max_row=func_end))
            chart.height = 7
            chart.width = 12
            sheet.add_chart(chart, "A17")
        if month_end >= 6:
            chart = LineChart()
            chart.title = "월별 운임 추이"
            chart.y_axis.title = "Loc Amt"
            chart.add_data(Reference(sheet, min_col=5, min_row=5, max_row=month_end), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=4, min_row=6, max_row=month_end))
            chart.height = 7
            chart.width = 12
            sheet.add_chart(chart, "G17")

        for column, width in {"A": 16, "B": 16, "D": 16, "E": 16, "G": 24, "H": 16, "J": 42, "K": 16}.items():
            sheet.column_dimensions[column].width = width

    @staticmethod
    def _write_overview_table(sheet, start_row: int, start_col: int, title: str, rows, style: ReportStyle) -> int:
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        border = Border(bottom=Side(style="thin", color="9EB6CE"))
        sheet.cell(start_row, start_col, title)
        sheet.cell(start_row, start_col + 1, "Loc Amt")
        for cell in (sheet.cell(start_row, start_col), sheet.cell(start_row, start_col + 1)):
            cell.font = Font(name="맑은 고딕", bold=True, color="1F2937")
            cell.fill = header_fill
            cell.border = border
        for offset, (label, amount) in enumerate(rows, start=1):
            row = start_row + offset
            sheet.cell(row, start_col, label)
            amount_cell = sheet.cell(row, start_col + 1, ReportWriter._numeric(amount))
            amount_cell.number_format = style.amount_format
        return start_row + len(rows)

    def _write_analytic_all(self, sheet, summaries: list[FuncCodeSummary], style: ReportStyle) -> None:
        self.formatter.apply_sheet_layout(sheet, style)
        sheet["A1"] = "통합 보고서"
        sheet["B1"] = ALL_SHEET_NAME
        self.formatter.style_row(sheet, 1, "title", style)
        self.formatter.style_row(sheet, 2, "blank", style)
        sheet["A3"] = "행 레이블"
        sheet["B3"] = "합계 : Loc Amt"
        self.formatter.style_row(sheet, 3, "header", style)

        month_lookup = {
            summary.func_code: {month.label: month for month in summary.months} for summary in summaries
        }
        months = sorted({month.label for summary in summaries for month in summary.months})
        row_index = 4
        for month_label in months:
            month_total = sum(
                (month_lookup[s.func_code][month_label].amount for s in summaries if month_label in month_lookup[s.func_code]),
                Decimal("0"),
            )
            sheet.cell(row_index, 1, month_label)
            sheet.cell(row_index, 2, self._numeric(month_total))
            self.formatter.style_row(sheet, row_index, "month", style)
            row_index += 1
            for summary in summaries:
                month = month_lookup[summary.func_code].get(month_label)
                if month is None:
                    continue
                sheet.cell(row_index, 1, summary.func_code)
                sheet.cell(row_index, 2, self._numeric(month.amount))
                self.formatter.style_row(sheet, row_index, "func", style)
                sheet.row_dimensions[row_index].outlineLevel = 1
                row_index += 1
                for port in month.ports:
                    sheet.cell(row_index, 1, port.name)
                    sheet.cell(row_index, 2, self._numeric(port.amount))
                    self.formatter.style_row(sheet, row_index, "analytic_port", style)
                    sheet.row_dimensions[row_index].outlineLevel = 2
                    row_index += 1
                    for customer in port.customers:
                        sheet.cell(row_index, 1, customer.name)
                        sheet.cell(row_index, 2, self._numeric(customer.amount))
                        self.formatter.style_row(sheet, row_index, "analytic_customer", style)
                        sheet.row_dimensions[row_index].outlineLevel = 3
                        row_index += 1

        sheet.cell(row_index, 1, "총합계")
        sheet.cell(row_index, 2, self._numeric(sum((s.amount for s in summaries), Decimal("0"))))
        self.formatter.style_row(sheet, row_index, "grand_total", style)

    def _write_raw_data(self, sheet, source_data: WorkbookData) -> None:
        sheet.append(source_data.headers)
        for row in source_data.rows:
            padded = list(row) + [None] * max(0, len(source_data.headers) - len(row))
            try:
                sheet.append(padded[: len(source_data.headers)])
            except IllegalCharacterError:
                sheet.append([self._safe_cell_value(value) for value in padded[: len(source_data.headers)]])

    def _write_summary_sheet(self, sheet, summary: FuncCodeSummary, style: ReportStyle) -> None:
        self.formatter.apply_sheet_layout(sheet, style)
        sheet["A1"] = "Func Code"
        sheet["B1"] = summary.func_code
        self.formatter.style_row(sheet, 1, "title", style)
        self.formatter.style_row(sheet, 2, "blank", style)
        sheet["A3"] = "행 레이블"
        sheet["B3"] = "합계 : Loc Amt"
        self.formatter.style_row(sheet, 3, "header", style)

        row_index = 4
        for month in summary.months:
            month_row = row_index
            sheet.cell(row=row_index, column=1, value=month.label)
            sheet.cell(row=row_index, column=2, value=self._numeric(month.amount))
            self.formatter.style_row(sheet, row_index, "month", style)
            row_index += 1
            for port in month.ports:
                port_row = row_index
                sheet.cell(row=row_index, column=1, value=port.name)
                sheet.cell(row=row_index, column=2, value=self._numeric(port.amount))
                self.formatter.style_row(sheet, row_index, "port", style)
                row_index += 1
                for customer in port.customers:
                    sheet.cell(row=row_index, column=1, value=customer.name)
                    sheet.cell(row=row_index, column=2, value=self._numeric(customer.amount))
                    self.formatter.style_row(sheet, row_index, "customer", style)
                    sheet.row_dimensions[row_index].outlineLevel = 2
                    row_index += 1
                sheet.row_dimensions[port_row].outlineLevel = 1

        sheet.cell(row=row_index, column=1, value="총합계")
        sheet.cell(row=row_index, column=2, value=self._numeric(summary.amount))
        self.formatter.style_row(sheet, row_index, "grand_total", style)

    @staticmethod
    def _numeric(amount: Decimal) -> int | float:
        if amount == amount.to_integral_value():
            return int(amount)
        return float(amount)

    @staticmethod
    def _safe_cell_value(value: Any) -> Any:
        if isinstance(value, str):
            return "".join(char for char in value if ord(char) >= 32 or char in "\t\n\r")
        return value
