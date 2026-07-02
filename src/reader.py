from __future__ import annotations

from io import BytesIO
import logging
from pathlib import Path
import re
from typing import Any
from xml.etree import ElementTree as ET
import zipfile

from openpyxl import load_workbook

from .config import BusinessRules, HEADER_SCAN_ROW_LIMIT
from .exceptions import MissingInputFileError, MissingWorksheetError
from .models import WorkbookData

SS_NS = "{urn:schemas-microsoft-com:office:spreadsheet}"
LOGGER = logging.getLogger(__name__)
BARE_AMPERSAND = re.compile(br"&(?!amp;|lt;|gt;|quot;|apos;|#\d+;|#x[0-9A-Fa-f]+;)")


class SourceWorkbookReader:
    def __init__(self, rules: BusinessRules | None = None) -> None:
        self.rules = rules or BusinessRules()

    def read(self, path: Path) -> WorkbookData:
        if not path.exists():
            raise MissingInputFileError(f"입력 파일을 찾을 수 없습니다: {path}")

        if self._is_excel_xml(path):
            try:
                sheet_name, header_row = self._find_excel_xml_source(path)
                return self._read_excel_xml(path, sheet_name, header_row)
            except ET.ParseError:
                repaired, replacement_count = self._repair_excel_xml(path)
                if replacement_count == 0:
                    raise
                LOGGER.warning("Repaired %s unescaped ampersands in %s", replacement_count, path)
                sheet_name, header_row = self._find_excel_xml_source(BytesIO(repaired))
                return self._read_excel_xml(BytesIO(repaired), sheet_name, header_row)
        if zipfile.is_zipfile(path):
            return self._read_xlsx(path)
        return self._read_xls(path)

    @staticmethod
    def _is_excel_xml(path: Path) -> bool:
        with path.open("rb") as file:
            return file.read(64).lstrip().startswith(b"<?xml")

    @staticmethod
    def _repair_excel_xml(path: Path) -> tuple[bytes, int]:
        return BARE_AMPERSAND.subn(b"&amp;", path.read_bytes())

    def _find_excel_xml_source(self, source) -> tuple[str, int]:
        candidates: list[tuple[str, int]] = []
        current_sheet = ""
        row_number = 0
        for event, elem in ET.iterparse(source, events=("start", "end")):
            if event == "start" and elem.tag.endswith("Worksheet"):
                current_sheet = elem.attrib.get(f"{SS_NS}Name", "")
                row_number = 0
            elif event == "end" and elem.tag.endswith("Row"):
                row_number += 1
                if row_number <= HEADER_SCAN_ROW_LIMIT:
                    values = self._xml_row_values(elem)
                    if self._is_header_row(values):
                        candidates.append((current_sheet, row_number))
                elem.clear()
        return self._select_candidate(candidates)

    def _read_excel_xml(self, source, sheet_name: str, header_row: int) -> WorkbookData:
        headers: list[str] | None = None
        rows: list[list[Any]] = []
        found_sheet = False
        in_target = False
        row_number = 0

        for event, elem in ET.iterparse(source, events=("start", "end")):
            if event == "start" and elem.tag.endswith("Worksheet"):
                current = elem.attrib.get(f"{SS_NS}Name")
                in_target = current == sheet_name
                found_sheet = found_sheet or in_target
            elif in_target and event == "end" and elem.tag.endswith("Row"):
                row_number += 1
                values = self._xml_row_values(elem)
                if row_number == header_row:
                    headers = ["" if value is None else str(value) for value in values]
                elif row_number > header_row:
                    rows.append(values)
                elem.clear()
            elif event == "end" and elem.tag.endswith("Row"):
                elem.clear()

        if not found_sheet or headers is None:
            raise MissingWorksheetError(f"원본 파일에 '{sheet_name}' 시트가 없습니다.")
        return WorkbookData(headers=headers, rows=rows, source_sheet_name=sheet_name)

    @staticmethod
    def _xml_row_values(row: ET.Element) -> list[Any]:
        values: list[Any] = []
        column_index = 1
        for cell in row:
            if not cell.tag.endswith("Cell"):
                continue
            explicit_index = cell.attrib.get(f"{SS_NS}Index")
            if explicit_index:
                while column_index < int(explicit_index):
                    values.append(None)
                    column_index += 1
            values.append(SourceWorkbookReader._xml_cell_value(cell))
            column_index += 1
        return values

    @staticmethod
    def _xml_cell_value(cell: ET.Element) -> Any:
        for child in cell:
            if child.tag.endswith("Data"):
                return child.text
        return None

    def _read_xlsx(self, path: Path) -> WorkbookData:
        workbook = load_workbook(path, data_only=True, read_only=True)
        candidates: list[tuple[str, int]] = []
        for worksheet in workbook.worksheets:
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROW_LIMIT, values_only=True), start=1
            ):
                if self._is_header_row(list(row)):
                    candidates.append((worksheet.title, row_number))
                    break
        sheet_name, header_row = self._select_candidate(candidates)
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(min_row=header_row, values_only=True)
        try:
            headers = ["" if value is None else str(value) for value in next(iterator)]
        except StopIteration as exc:
            raise MissingWorksheetError(f"'{sheet_name}' 시트가 비어 있습니다.") from exc
        return WorkbookData(headers=headers, rows=[list(row) for row in iterator], source_sheet_name=sheet_name)

    def _read_xls(self, path: Path) -> WorkbookData:
        try:
            import xlrd
        except ImportError as exc:
            raise MissingInputFileError("구형 .xls 파일을 읽으려면 xlrd가 필요합니다.") from exc

        workbook = xlrd.open_workbook(str(path))
        candidates: list[tuple[str, int]] = []
        for name in workbook.sheet_names():
            candidate_sheet = workbook.sheet_by_name(name)
            for row_index in range(min(candidate_sheet.nrows, HEADER_SCAN_ROW_LIMIT)):
                if self._is_header_row(candidate_sheet.row_values(row_index)):
                    candidates.append((name, row_index + 1))
                    break
        sheet_name, header_row = self._select_candidate(candidates)
        sheet = workbook.sheet_by_name(sheet_name)
        if sheet.nrows == 0:
            raise MissingWorksheetError(f"'{sheet_name}' 시트가 비어 있습니다.")
        header_index = header_row - 1
        headers = ["" if value is None else str(value) for value in sheet.row_values(header_index)]
        rows = [sheet.row_values(index) for index in range(header_index + 1, sheet.nrows)]
        return WorkbookData(headers=headers, rows=rows, source_sheet_name=sheet_name)

    def _is_header_row(self, values: list[Any]) -> bool:
        normalized = {self._normalize_header(value) for value in values if value not in (None, "")}
        required = {self._normalize_header(value) for value in self.rules.required_columns}
        return required.issubset(normalized)

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return " ".join(str(value).replace("\ufeff", "").split()).casefold()

    @staticmethod
    def _select_candidate(candidates: list[tuple[str, int]]) -> tuple[str, int]:
        if not candidates:
            raise MissingWorksheetError(
                "필수 컬럼을 모두 포함한 Source 시트를 찾지 못했습니다. "
                "필요 컬럼: Job Date, Func Code, Arrival Port, Depart Port, Customer Name, Loc Amt"
            )
        if len(candidates) > 1:
            names = ", ".join(f"{name}({row}행)" for name, row in candidates)
            raise MissingWorksheetError(f"Source 후보 시트가 여러 개입니다: {names}")
        return candidates[0]
